"""generate_company_list.py — Print all unique company names for industry categorization."""
from pathlib import Path
import openpyxl, config

def main():
    xl = Path(config.EXCEL_PATH)
    if not xl.exists():
        print(f"ERROR: Workbook not found at {xl}"); return
    wb = openpyxl.load_workbook(xl, read_only=True)
    companies = set()
    for sheet_name, col_name in [("Employment History","company"),("Current Snapshot","current_company")]:
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try: ci = headers.index(col_name)
        except ValueError: continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[ci] and str(row[ci]).strip(): companies.add(str(row[ci]).strip())
    wb.close()
    sorted_cos = sorted(companies)
    print(f"\nFound {len(sorted_cos)} unique companies:\n")
    print(", ".join(sorted_cos))
    print('\n\nSend to Claude: "Categorize by industry, return {\"Company\": \"Industry\"} JSON."')

if __name__ == "__main__": main()
