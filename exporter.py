"""exporter.py — Excel workbook helper utilities."""
from __future__ import annotations
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

LINKEDIN_BLUE = "0A66C2"
HEADER_FILL   = PatternFill("solid", fgColor=LINKEDIN_BLUE)
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL      = PatternFill("solid", fgColor="F3F6FA")

def style_header_row(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

def get_header_map(ws) -> dict[str, int]:
    return {cell.value: cell.column for cell in ws[1] if cell.value}

def ensure_column(ws, col_name: str) -> int:
    hmap = get_header_map(ws)
    if col_name in hmap:
        return hmap[col_name]
    new_idx = ws.max_column + 1
    cell = ws.cell(row=1, column=new_idx, value=col_name)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    return new_idx

def append_row(ws, headers: list[str], data: dict):
    row_num = ws.max_row + 1
    fill = ALT_FILL if row_num % 2 == 0 else None
    for i, col in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=i, value=data.get(col))
        if fill: cell.fill = fill

def auto_width(ws, min_width=10, max_width=40):
    for col in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(min_width, min(max_width, length + 2))
