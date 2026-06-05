"""
enrich_to_excel.py  —  Bridge: enriched_profiles.json → network_master.xlsx
=============================================================================
Reads every fully-enriched profile from enriched_profiles.json and writes
the data back into the matching row of network_master.xlsx (Current Snapshot
+ Employment History sheets).

Run this after enricher.py completes, then run export_to_json.py to push
the updated data to the front-end.

Typical daily workflow:
    python enricher.py           ← fetches 8 profiles, saves to enriched_profiles.json
    python enrich_to_excel.py    ← THIS FILE: syncs enriched_profiles.json → Excel
    python export_to_json.py     ← rebuilds network.json from Excel
    python git_push.py           ← pushes docs/ to GitHub Pages

What this script writes
-----------------------
Current Snapshot sheet (one row per connection):
    headline, location, profile_photo_url, banner_url
    last_enriched, enriched_version
    greek_orgs          pipe-separated display names   e.g. "Pi Beta Phi|Theta Chi"
    greek_org_ids       pipe-separated org ids         e.g. "piphi|thetachi"  [for front-end]
    greek_org_names     same as greek_orgs (alias)
    cert_tags           pipe-separated cert tags        e.g. "PE|LEED"
    skills              semicolon-separated             e.g. "Python;AutoCAD"
    education           semicolon-separated school entries (for export_to_json)
    honors              pipe-separated honor titles
    org_memberships     pipe-separated org names
    summary             (headline used as summary if no separate field)

Employment History sheet (one row per job, appended if not already present):
    connection_id, company, title, start_date, end_date, is_current
    (change_detector.py owns this sheet — we only ADD rows, never modify existing)

Columns added to Excel if missing:
    greek_orgs, greek_org_ids, greek_org_names, cert_tags,
    skills, education, honors, org_memberships, banner_url

Greek org ID mapping
--------------------
org names from html_parser → matched against GREEK_NAME_TO_ID lookup table
→ written as pipe-separated IDs to greek_org_ids column
→ front-end reads this as c.greek_org_ids for auto-badge display
"""

from __future__ import annotations

import json
import logging
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────────────
# Enricher-v1 data (local, never on GitHub)
ENRICHED_FILE = Path(r"B:\linkedin-data\enriched_profiles.json")

# Pipeline Excel (local, never on GitHub)
EXCEL_PATH    = Path(r"B:\linkedin-data\network_master.xlsx")

# ── Logging ───────────────────────────────────────────────────────────────────
LOG_DIR = Path(r"B:\linkedin-data\logs")
LOG_DIR.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  [BRIDGE]  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / f"bridge_{date.today()}.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("bridge")

# ── Greek org name → front-end ID mapping ─────────────────────────────────────
# Maps display names (as detected by html_parser) to the id strings
# used in GREEK_DB_ALL in index.html so greek_org_ids auto-applies badges.
GREEK_NAME_TO_ID: dict[str, str] = {
    # Fraternities
    "pi kappa alpha":        "pike",
    "pike":                  "pike",
    "phi gamma delta":       "fiji",
    "fiji":                  "fiji",
    "sigma phi epsilon":     "sigep",
    "sig ep":                "sigep",
    "sigma chi":             "sigchi",
    "sig chi":               "sigchi",
    "sigma alpha epsilon":   "sae",
    "phi delta theta":       "phidelt",
    "phi delt":              "phidelt",
    "beta theta pi":         "beta",
    "kappa sigma":           "kapsig",
    "kappa sig":             "kapsig",
    "lambda chi alpha":      "lambdachi",
    "lambda chi":            "lambdachi",
    "theta chi":             "thetachi",
    "delta tau delta":       "delt",
    "sigma nu":              "sigmanu",
    "sig nu":                "sigmanu",
    "alpha tau omega":       "ato",
    "tau kappa epsilon":     "tke",
    "phi kappa psi":         "phipsi",
    "phi psi":               "phipsi",
    "delta kappa epsilon":   "dke",
    "deke":                  "dke",
    "pi kappa phi":          "pikapp",
    "alpha epsilon pi":      "aepi",
    "alpha phi alpha":       "apa",
    "kappa alpha psi":       "kapsi",
    "omega psi phi":         "omegaps",
    "phi beta sigma":        "phibets",
    # Sororities
    "alpha delta pi":        "adpi",
    "adpi":                  "adpi",
    "alpha epsilon phi":     "aephi",
    "alpha phi":             "aphi",
    "a phi":                 "aphi",
    "alpha chi omega":       "axo",
    "alpha chi":             "axo",
    "delta gamma":           "dg",
    "delta delta delta":     "tridelt",
    "tri delta":             "tridelt",
    "tri delt":              "tridelt",
    "delta zeta":            "dz",
    "gamma phi beta":        "gammaphi",
    "gamma phi":             "gammaphi",
    "kappa alpha theta":     "kat",
    "kappa delta":           "kd",
    "kappa kappa gamma":     "kkg",
    "pi beta phi":           "piphi",
    "pi phi":                "piphi",
    "zeta tau alpha":        "zta",
    "zeta":                  "zta",
    "sigma kappa":           "sigmakap",
    "sig kap":               "sigmakap",
    "chi omega":             "chiomega",
    "chi o":                 "chiomega",
    "delta sigma theta":     "dst",
    "deltas":                "dst",
}


def _org_name_to_id(name: str) -> str | None:
    """Map a detected org name to its front-end ID. Case-insensitive."""
    return GREEK_NAME_TO_ID.get(name.strip().lower())


# ── Date parsing helpers ───────────────────────────────────────────────────────

_MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

def _parse_date_range(date_range: str | None) -> tuple[int | None, int | None, int | None, int | None]:
    """
    Parse a date range string like 'Jan 2022 - Mar 2024' or 'Sep 2023 - Present'.
    Returns (start_month, start_year, end_month, end_year).
    end_month/end_year are None for 'Present'.
    """
    if not date_range:
        return None, None, None, None

    # Strip duration suffix: "Jan 2022 - Mar 2024 · 2 yrs"
    date_range = date_range.split("·")[0].strip()

    parts = re.split(r"\s*[-–]\s*", date_range, maxsplit=1)
    start_str = parts[0].strip() if parts else ""
    end_str   = parts[1].strip() if len(parts) > 1 else ""

    def parse_one(s: str) -> tuple[int | None, int | None]:
        s = s.strip().lower()
        if not s or s == "present":
            return None, None
        m = re.search(r"(\w+)\s+(\d{4})", s)
        if m:
            month = _MONTH_MAP.get(m.group(1)[:3])
            year  = int(m.group(2))
            return month, year
        # Year only
        m2 = re.search(r"(\d{4})", s)
        if m2:
            return None, int(m2.group(1))
        return None, None

    sm, sy = parse_one(start_str)
    em, ey = parse_one(end_str)
    return sm, sy, em, ey


def _is_current(date_range: str | None) -> bool:
    if not date_range:
        return False
    return "present" in date_range.lower()


# ── Education string builder ───────────────────────────────────────────────────

def _build_edu_string(education: list[dict]) -> str:
    """
    Convert education list → semicolon-separated string for Excel.
    Format matches what export_to_json.py's parse_edu_string() expects:
        School Name (Degree, Field, YYYY-YYYY); ...
    """
    parts = []
    for edu in education:
        school = edu.get("school") or ""
        if not school or school.lower() in ("nothing to see for now", ""):
            continue
        detail_parts = []
        degree = edu.get("degree") or ""
        field  = edu.get("field")  or ""
        if degree:
            detail_parts.append(degree)
        if field:
            detail_parts.append(field)
        # Parse years from date_range
        dr = edu.get("date_range") or ""
        year_m = re.search(r"(\d{4})\s*[-–]\s*(\d{4})", dr)
        if year_m:
            detail_parts.append(f"{year_m.group(1)}-{year_m.group(2)}")
        if detail_parts:
            parts.append(f"{school} ({', '.join(detail_parts)})")
        else:
            parts.append(school)
    return "; ".join(parts)


# ── Greek org field builders ───────────────────────────────────────────────────

def _build_greek_fields(greek_orgs: list[str]) -> tuple[str, str, str]:
    """
    Returns (greek_orgs_pipe, greek_org_ids_pipe, greek_org_names_pipe).
    Filters out empty-state placeholder text.
    """
    clean = [
        g for g in greek_orgs
        if g and "nothing to see" not in g.lower()
    ]
    if not clean:
        return "", "", ""

    ids = []
    for name in clean:
        org_id = _org_name_to_id(name)
        if org_id:
            ids.append(org_id)

    return (
        "|".join(clean),
        ",".join(ids),   # comma-separated for front-end greek_org_ids field
        "|".join(clean),
    )


# ── Experience → Employment History rows ──────────────────────────────────────

def _experience_to_history_rows(profile: dict) -> list[dict]:
    """
    Flatten the enricher's nested experience structure into flat rows
    suitable for the Employment History sheet.

    Enricher experience structure:
        {company, company_location, total_duration, roles:[{title, date_range, ...}]}
    """
    rows = []
    connection_id = profile["id"]

    for exp in profile.get("experience", []):
        company = exp.get("company") or ""
        if not company or "nothing to see" in company.lower():
            continue

        roles = exp.get("roles") or []
        if not roles:
            # No role detail — write one row with company only
            rows.append({
                "connection_id": connection_id,
                "company":       company,
                "title":         "",
                "start_month":   None,
                "start_year":    None,
                "end_month":     None,
                "end_year":      None,
                "is_current":    False,
            })
            continue

        for role in roles:
            title      = role.get("title") or ""
            date_range = role.get("date_range")
            sm, sy, em, ey = _parse_date_range(date_range)
            current    = _is_current(date_range)

            rows.append({
                "connection_id": connection_id,
                "company":       company,
                "title":         title,
                "start_month":   sm,
                "start_year":    sy,
                "end_month":     None if current else em,
                "end_year":      None if current else ey,
                "is_current":    current,
            })

    return rows


# ── Excel helpers ──────────────────────────────────────────────────────────────

def _ensure_column(ws, headers: list, col_name: str) -> int:
    """Add a column header if missing. Returns 1-indexed column number."""
    if col_name in headers:
        return headers.index(col_name) + 1
    next_col = len(headers) + 1
    ws.cell(row=1, column=next_col, value=col_name)
    headers.append(col_name)
    return next_col


def _get_headers(ws) -> list[str]:
    return [
        (str(c.value).strip() if c.value is not None else f"__col_{i}")
        for i, c in enumerate(ws[1])
    ]


def _find_row_for_id(ws, id_col: int, target_id: str) -> int | None:
    for row in ws.iter_rows(min_row=2):
        cell_val = row[id_col - 1].value
        if cell_val is not None and str(cell_val).strip() == target_id:
            return row[0].row
    return None


# ── Main sync logic ────────────────────────────────────────────────────────────

def sync():
    # ── Validate inputs ──────────────────────────────────────────────────────
    if not ENRICHED_FILE.exists():
        log.error("enriched_profiles.json not found at %s", ENRICHED_FILE)
        log.error("Run enricher.py first.")
        sys.exit(1)

    if not EXCEL_PATH.exists():
        log.error("network_master.xlsx not found at %s", EXCEL_PATH)
        log.error("Run pipeline.py first.")
        sys.exit(1)

    # ── Load enriched profiles ───────────────────────────────────────────────
    log.info("Loading enriched profiles: %s", ENRICHED_FILE)
    with open(ENRICHED_FILE, encoding="utf-8") as f:
        enriched: dict = json.load(f)

    total        = len(enriched)
    full_count   = sum(1 for p in enriched.values() if p.get("enrichment_status") == "full")
    partial_count = sum(1 for p in enriched.values() if p.get("enrichment_status") == "partial")
    log.info("Loaded %d profiles (%d full, %d partial)", total, full_count, partial_count)

    # Only sync profiles that have actual data worth writing
    to_sync = [
        p for p in enriched.values()
        if p.get("enrichment_status") in ("full", "partial")
        and p.get("enriched_on")
    ]
    log.info("Profiles to sync into Excel: %d", len(to_sync))

    if not to_sync:
        log.info("Nothing to sync — run enricher.py first.")
        return

    # ── Open workbook ────────────────────────────────────────────────────────
    log.info("Opening workbook: %s", EXCEL_PATH)
    wb = openpyxl.load_workbook(EXCEL_PATH)

    if "Current Snapshot" not in wb.sheetnames:
        log.error("'Current Snapshot' sheet not found in workbook.")
        sys.exit(1)

    snap_ws = wb["Current Snapshot"]
    snap_headers = _get_headers(snap_ws)

    # Ensure all required columns exist
    NEW_SNAP_COLS = [
        "greek_orgs", "greek_org_ids", "greek_org_names",
        "cert_tags", "skills", "education", "honors",
        "org_memberships", "banner_url", "headline",
        "profile_photo_url", "last_enriched", "enriched_version",
    ]
    for col in NEW_SNAP_COLS:
        _ensure_column(snap_ws, snap_headers, col)

    # Get / ensure Employment History sheet
    if "Employment History" not in wb.sheetnames:
        hist_ws = wb.create_sheet("Employment History")
        hist_ws.append([
            "connection_id", "company", "title",
            "start_month", "start_year", "end_month", "end_year", "is_current",
        ])
        log.info("Created 'Employment History' sheet")
    else:
        hist_ws = wb["Employment History"]

    hist_headers = _get_headers(hist_ws)

    # Build a set of existing history rows to avoid duplicates:
    # key = (connection_id, company, title)
    existing_history: set[tuple] = set()
    cid_col_h = hist_headers.index("connection_id") + 1 if "connection_id" in hist_headers else None
    co_col_h  = hist_headers.index("company") + 1      if "company"        in hist_headers else None
    ti_col_h  = hist_headers.index("title") + 1        if "title"          in hist_headers else None

    if cid_col_h and co_col_h and ti_col_h:
        for row in hist_ws.iter_rows(min_row=2, values_only=True):
            key = (
                str(row[cid_col_h - 1] or "").strip(),
                str(row[co_col_h  - 1] or "").strip(),
                str(row[ti_col_h  - 1] or "").strip(),
            )
            existing_history.add(key)

    # ── Per-profile sync ─────────────────────────────────────────────────────
    snap_updated = 0
    snap_skipped = 0
    hist_added   = 0

    # Column index helpers for snapshot
    def sc(name):
        return snap_headers.index(name) + 1 if name in snap_headers else None

    id_col = sc("connection_id") or sc("linkedin_id")
    if not id_col:
        log.error("Cannot find connection_id column in Current Snapshot.")
        sys.exit(1)

    for profile in to_sync:
        profile_id = str(profile.get("id", "")).strip()
        if not profile_id:
            continue

        # ── Find matching row in snapshot ────────────────────────────────────
        target_row = _find_row_for_id(snap_ws, id_col, profile_id)

        if not target_row:
            log.warning("Profile %s not found in snapshot — skipping", profile_id)
            snap_skipped += 1
            continue

        def wc(col_name: str, value) -> None:
            idx = sc(col_name)
            if idx:
                snap_ws.cell(row=target_row, column=idx, value=value)

        # ── Build Greek fields ───────────────────────────────────────────────
        greek_pipe, greek_ids, greek_names = _build_greek_fields(
            profile.get("greek_orgs", [])
        )

        # ── Build cert_tags string ───────────────────────────────────────────
        cert_tags_pipe = "|".join(
            t for t in profile.get("cert_tags", [])
            if t and "nothing to see" not in t.lower()
        )

        # ── Build skills string ──────────────────────────────────────────────
        skills_semi = ";".join(
            s for s in profile.get("skills", [])
            if s and len(s) > 1
        )

        # ── Build education string ───────────────────────────────────────────
        edu_str = _build_edu_string(profile.get("education", []))

        # ── Build honors string ──────────────────────────────────────────────
        honors_pipe = "|".join(
            h.get("title", "") for h in profile.get("honors", [])
            if h.get("title") and "nothing to see" not in h["title"].lower()
        )

        # ── Build org_memberships string ─────────────────────────────────────
        orgs_pipe = "|".join(
            o.get("name", "") for o in profile.get("organizations", [])
            if o.get("name") and "nothing to see" not in o["name"].lower()
        )

        # ── Write to snapshot ────────────────────────────────────────────────
        wc("headline",          profile.get("headline", ""))
        wc("location",          profile.get("location", ""))
        wc("profile_photo_url", profile.get("photo_url", ""))
        wc("banner_url",        profile.get("banner_url", ""))
        wc("last_enriched",     profile.get("enriched_on", ""))
        wc("enriched_version",  "3.0")
        wc("greek_orgs",        greek_pipe)
        wc("greek_org_ids",     greek_ids)
        wc("greek_org_names",   greek_names)
        wc("cert_tags",         cert_tags_pipe)
        wc("skills",            skills_semi)
        wc("education",         edu_str)
        wc("honors",            honors_pipe)
        wc("org_memberships",   orgs_pipe)

        snap_updated += 1

        # ── Write employment history rows ─────────────────────────────────────
        history_rows = _experience_to_history_rows(profile)
        for hr in history_rows:
            key = (
                str(hr["connection_id"]).strip(),
                str(hr["company"]).strip(),
                str(hr["title"]).strip(),
            )
            if key in existing_history:
                continue  # already in sheet — don't duplicate

            hist_ws.append([
                hr["connection_id"],
                hr["company"],
                hr["title"],
                hr["start_month"],
                hr["start_year"],
                hr["end_month"],
                hr["end_year"],
                "YES" if hr["is_current"] else "NO",
            ])
            existing_history.add(key)
            hist_added += 1

    # ── Save ─────────────────────────────────────────────────────────────────
    log.info("Saving workbook...")
    try:
        wb.save(EXCEL_PATH)
        log.info("Saved successfully.")
    except Exception as exc:
        log.error("Failed to save workbook: %s", exc)
        sys.exit(1)
    finally:
        wb.close()

    # ── Summary ──────────────────────────────────────────────────────────────
    log.info("=" * 55)
    log.info("Sync complete.")
    log.info("  Snapshot rows updated : %d", snap_updated)
    log.info("  Snapshot rows skipped : %d (not in Excel)", snap_skipped)
    log.info("  History rows added    : %d", hist_added)
    log.info("=" * 55)
    log.info("Next step: python export_to_json.py")


if __name__ == "__main__":
    sync()
