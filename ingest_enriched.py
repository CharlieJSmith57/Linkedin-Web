"""
ingest_enriched.py — Scraper data ingestion bridge

Reads enriched_profiles.json produced by the LinkedIn HTML scraper
(linkedin-enricher-v1) and merges its data into network_master.xlsx.

The scraper lives at:  B:\\linkedin-network\\network-master-csv\\enriched_profiles.json
The master workbook:   B:\\linkedin-data\\network_master.xlsx

Run manually after the enricher has processed new profiles:
    python ingest_enriched.py

Or run automatically — add it to pipeline.py's daily cycle.

MERGE STRATEGY
──────────────
- Matches profiles by linkedin_public_id (the URL slug, e.g. "jane-smith-123")
- Non-empty scraper fields OVERWRITE existing workbook fields
- Empty scraper fields PRESERVE existing workbook values
  (so a partial scrape doesn't erase data from a previous good fetch)
- New fields added by the scraper that don't exist in the workbook
  get new columns created automatically

ENRICHED FIELDS INGESTED
─────────────────────────
From experience entries:  company_location, employment_type, skills_used
From education entries:   activities, greek_org
From certifications:      cert tags (is_pe, is_cxa, is_bcxp), credential_id
From organizations:       org_name, org_role, org_dates, is_greek
From honors:              honor_name, issuer, date
From profile top level:   headline, location, photo_url, banner_url
Aggregated:               all_cert_tags, all_greek_orgs, all_skills (expanded)
"""

from __future__ import annotations

import json
import logging
from datetime import date
from pathlib import Path

import openpyxl

import config
from exporter import ensure_column, get_header_map

log = logging.getLogger(__name__)

# Path to the scraper's output file
ENRICHED_JSON_PATH = Path(r"B:\linkedin-network\network-master-csv\enriched_profiles.json")

EXCEL_PATH = Path(config.EXCEL_PATH)


# ── Field mapping ─────────────────────────────────────────────────────────────
# Maps scraper field names → Excel column names added by this module.
# These are written into the Current Snapshot sheet.

SCALAR_FIELDS = {
    # Scraper key              → Excel column name
    "headline":                "headline",
    "location":                "location",
    "profile_photo":           "profile_photo_url",   # top-level photo record
    "banner":                  "banner_url",
    "summary":                 "summary",
}

# Aggregated fields we compute from nested scraper data
COMPUTED_COLUMNS = [
    "cert_tags",        # pipe-separated: "PE|CxA"
    "greek_orgs",       # pipe-separated org names
    "company_locations",# pipe-separated: "Company::City, State"
    "honors",           # pipe-separated: "Award Name (Issuer, Year)"
    "org_memberships",  # pipe-separated org memberships
    "enricher_source",  # "html_scraper" — marks which rows came from scraper
    "scraper_synced_on",# ISO date of last sync from scraper
]


# ── Load scraper data ─────────────────────────────────────────────────────────

def load_enriched_profiles() -> dict:
    """
    Load enriched_profiles.json from the scraper's data directory.
    Returns {linkedin_public_id: profile_dict}.
    """
    if not ENRICHED_JSON_PATH.exists():
        log.warning("enriched_profiles.json not found at %s", ENRICHED_JSON_PATH)
        log.warning("Run the linkedin-enricher-v1 scraper first.")
        return {}

    with open(ENRICHED_JSON_PATH, encoding="utf-8") as f:
        raw = json.load(f)

    # enriched_profiles.json structure:
    # { "linkedin_public_id": { ...profile data... }, ... }
    # OR a list of profiles with a "public_id" field — handle both
    if isinstance(raw, dict):
        # Remove metadata keys that start with _
        return {k: v for k, v in raw.items() if not k.startswith("_")}
    elif isinstance(raw, list):
        result = {}
        for p in raw:
            pid = p.get("public_id") or p.get("linkedin_public_id") or p.get("id")
            if pid:
                result[pid] = p
        return result
    else:
        log.error("Unexpected enriched_profiles.json format")
        return {}


# ── Extract aggregated fields from a profile ──────────────────────────────────

def extract_cert_tags(profile: dict) -> str:
    """Return pipe-separated cert tags: 'PE|CxA|BCxP'"""
    tags = set()
    for cert in profile.get("certifications", []):
        if cert.get("is_pe"):    tags.add("PE")
        if cert.get("is_cxa"):   tags.add("CxA")
        if cert.get("is_bcxp"):  tags.add("BCxP")
    # Also check top-level flags if scraper set them
    if profile.get("is_pe"):   tags.add("PE")
    if profile.get("is_cxa"):  tags.add("CxA")
    if profile.get("is_bcxp"): tags.add("BCxP")
    return "|".join(sorted(tags))


def extract_greek_orgs(profile: dict) -> str:
    """Return pipe-separated greek org names found anywhere in the profile."""
    orgs = set()
    # From organizations section
    for org in profile.get("organizations", []):
        if org.get("is_greek") and org.get("name"):
            orgs.add(org["name"])
    # From education activities
    for edu in profile.get("education", []):
        go = edu.get("greek_org")
        if go:
            orgs.add(go)
    # From experience (some people list their fraternity as an employer)
    for exp in profile.get("experience", []):
        if exp.get("is_greek") and exp.get("company"):
            orgs.add(exp["company"])
    return "|".join(sorted(orgs))


def extract_company_locations(profile: dict) -> str:
    """Return pipe-separated 'Company::Location' for all experience entries."""
    locs = []
    for exp in profile.get("experience", []):
        co  = exp.get("company", "")
        loc = exp.get("location", "")
        if co and loc:
            locs.append(f"{co}::{loc}")
    return "|".join(locs)


def extract_honors(profile: dict) -> str:
    """Return pipe-separated honor strings."""
    items = []
    for h in profile.get("honors", []):
        name   = h.get("name", "")
        issuer = h.get("issuer", "")
        yr     = h.get("year", "") or h.get("date", "")
        parts  = [name]
        if issuer: parts.append(issuer)
        if yr:     parts.append(str(yr))
        items.append(" · ".join(parts))
    return "|".join(items)


def extract_org_memberships(profile: dict) -> str:
    """Return pipe-separated org membership strings."""
    items = []
    for org in profile.get("organizations", []):
        name = org.get("name", "")
        role = org.get("role", "")
        if name:
            items.append(f"{name}: {role}" if role else name)
    return "|".join(items)


def extract_photo_url(profile: dict) -> str:
    """Extract profile photo URL or abs_path from the profile photo record."""
    photo = profile.get("profile_photo")
    if not photo:
        return profile.get("photo_url", "") or ""
    if isinstance(photo, dict):
        # Prefer abs_path (local disk), fall back to src_raw
        return photo.get("abs_path", "") or photo.get("src_raw", "")
    return str(photo)


def extract_banner_url(profile: dict) -> str:
    """Extract banner URL from the banner record."""
    banner = profile.get("banner")
    if not banner:
        return ""
    if isinstance(banner, dict):
        return banner.get("abs_path", "") or banner.get("src_raw", "")
    return str(banner)


# ── Workbook operations ───────────────────────────────────────────────────────

def load_snapshot_index(ws) -> dict[str, int]:
    """Return {linkedin_public_id: row_number} for the snapshot sheet."""
    headers = get_header_map(ws)
    pid_col = headers.get("linkedin_public_id")
    if not pid_col:
        log.warning("No 'linkedin_public_id' column found in snapshot sheet")
        return {}

    index = {}
    for row in ws.iter_rows(min_row=2):
        val = row[pid_col - 1].value
        if val:
            index[str(val).strip()] = row[0].row
    return index


def write_enriched_row(ws, row_num: int, col_map: dict[str, int], data: dict):
    """Write enriched fields into a specific row, skipping empty values."""
    for field, col_idx in col_map.items():
        value = data.get(field)
        if value is None or value == "":
            continue  # preserve existing value
        ws.cell(row=row_num, column=col_idx, value=value)


# ── Main ingest ───────────────────────────────────────────────────────────────

def ingest():
    log.info("=== Scraper data ingestion starting ===")

    # Load scraper data
    enriched = load_enriched_profiles()
    if not enriched:
        log.info("No enriched profiles to ingest. Done.")
        return

    log.info("Loaded %d enriched profiles from scraper", len(enriched))

    # Load workbook
    if not EXCEL_PATH.exists():
        log.error("network_master.xlsx not found at %s", EXCEL_PATH)
        log.error("Run pipeline.py at least once first to create it.")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Current Snapshot"]

    # Ensure all enriched columns exist
    all_new_cols = list(COMPUTED_COLUMNS) + ["banner_url"]
    col_map = {col: ensure_column(ws, col) for col in all_new_cols}

    # Load existing row index
    snap_index = load_snapshot_index(ws)
    log.info("Snapshot has %d rows", len(snap_index))

    headers = get_header_map(ws)
    matched = 0
    unmatched = []

    today = date.today().isoformat()

    for public_id, profile in enriched.items():
        row_num = snap_index.get(public_id)
        if row_num is None:
            # Try stripping URL prefix if scraper stored full URL
            slug = public_id.split("/in/")[-1].strip("/")
            row_num = snap_index.get(slug)

        if row_num is None:
            unmatched.append(public_id)
            continue

        # Build data dict for this row
        data = {
            "cert_tags":         extract_cert_tags(profile),
            "greek_orgs":        extract_greek_orgs(profile),
            "company_locations": extract_company_locations(profile),
            "honors":            extract_honors(profile),
            "org_memberships":   extract_org_memberships(profile),
            "enricher_source":   "html_scraper",
            "scraper_synced_on": today,
            "banner_url":        extract_banner_url(profile),
        }

        # Scalar fields — only write if non-empty and column exists
        for scraper_key, col_name in SCALAR_FIELDS.items():
            col_idx = headers.get(col_name)
            if not col_idx:
                continue
            if scraper_key == "profile_photo":
                val = extract_photo_url(profile)
            elif scraper_key == "banner":
                val = extract_banner_url(profile)
            else:
                val = profile.get(scraper_key, "")
            if val:
                data[col_name] = val
                col_map[col_name] = col_idx

        write_enriched_row(ws, row_num, col_map, data)
        matched += 1

    wb.save(EXCEL_PATH)

    log.info("Ingested: %d profiles matched and updated", matched)
    if unmatched:
        log.warning(
            "%d profiles in enriched_profiles.json had no match in the workbook: %s",
            len(unmatched),
            unmatched[:10],  # show first 10
        )

    log.info("Workbook saved: %s", EXCEL_PATH)
    log.info("=== Ingestion complete ===")


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [INGEST] %(levelname)s %(message)s",
    )
    ingest()
