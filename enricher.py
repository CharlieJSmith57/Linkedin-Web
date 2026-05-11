"""
enricher.py — Slow-crawl background enricher
Runs independently of the daily CSV pipeline.
Fetches 8–12 random profiles per day with randomized timing
to minimize LinkedIn detection risk.

NEW FIELDS ADDED (on top of prior version):
  - education        list of {school, degree, field, start_year, end_year}
  - school           shorthand: the most recent / primary school name
  - title_standard   normalized seniority bucket derived from raw title
  - industry         looked up from industry_map.json (you populate this)
  - skills           list of skill strings from the LinkedIn profile
"""

import json
import random
import time
import logging
import os
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl
from linkedin_api import Linkedin

# ── Config ──────────────────────────────────────────────────────────────────
import config  # config.py lives alongside this file
from title_taxonomy import classify as classify_title

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [ENRICHER] %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

# How many profiles to fetch in one daily run (chosen randomly within this range)
DAILY_MIN = 8
DAILY_MAX = 12

# Seconds to wait between individual profile fetches (randomized within range)
DELAY_MIN = 45   # never faster than 45 s between requests
DELAY_MAX = 180  # up to 3 min — looks like a human browsing slowly

# Path to the industry mapping file (JSON dict: {"Google": "Technology", ...})
INDUSTRY_MAP_PATH = Path(config.DATA_DIR) / "industry_map.json"

# Path to the master Excel workbook written by exporter.py
EXCEL_PATH = Path(config.EXCEL_PATH)


# ── Seniority normalization ──────────────────────────────────────────────────
# Maps raw job title keywords → a standardized seniority bucket.
# Order matters: checked top-to-bottom, first match wins.
SENIORITY_RULES = [
    ("C-Suite",       ["ceo", "coo", "cfo", "cto", "cmo", "chief", "president", "founder"]),
    ("VP",            ["vice president", "vp ", "svp", "evp", "group vp"]),
    ("Director",      ["director", "managing director", "md,"]),
    ("Principal",     ["principal", "staff engineer", "distinguished"]),
    ("Manager",       ["manager", "head of", "lead,", " lead ", "team lead"]),
    ("Senior",        ["senior", "sr.", "sr ", "level 5", "level 6", "l5", "l6"]),
    ("Mid-level",     ["engineer", "analyst", "designer", "developer", "consultant",
                       "scientist", "associate", "specialist", "architect"]),
    ("Entry-level",   ["junior", "jr.", "entry", "associate engineer", "coordinator"]),
    ("Intern",        ["intern", "internship", "co-op", "coop"]),
]

def standardize_title(raw_title: str) -> str:
    """Return a normalized seniority bucket for a raw LinkedIn job title."""
    if not raw_title:
        return "Unknown"
    t = raw_title.lower()
    for bucket, keywords in SENIORITY_RULES:
        if any(k in t for k in keywords):
            return bucket
    return "Other"


# ── Industry lookup ──────────────────────────────────────────────────────────
def load_industry_map() -> dict:
    """
    Load the industry map from industry_map.json.
    Returns empty dict if the file doesn't exist yet — you build it over time
    by running: python generate_company_list.py   (see that file)
    """
    if INDUSTRY_MAP_PATH.exists():
        with open(INDUSTRY_MAP_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def lookup_industry(company_name: str, industry_map: dict) -> str:
    """
    Exact match first, then case-insensitive, then 'Unknown'.
    You can enrich this file any time without re-running the enricher.
    """
    if not company_name:
        return "Unknown"
    if company_name in industry_map:
        return industry_map[company_name]
    lower = company_name.lower()
    for k, v in industry_map.items():
        if k.lower() == lower:
            return v
    return "Unknown"


# ── Education parsing ────────────────────────────────────────────────────────
def parse_education(raw_profile: dict) -> list[dict]:
    """
    Extract education entries from a raw linkedin-api profile dict.
    Returns a list of dicts, most recent first:
      {school, degree, field, start_year, end_year}
    """
    entries = []
    for edu in raw_profile.get("education", []):
        school = ""
        # School name lives at different keys depending on API version
        if "schoolName" in edu:
            school = edu["schoolName"]
        elif "school" in edu and isinstance(edu["school"], dict):
            school = edu["school"].get("schoolName", "")

        degree = edu.get("degreeName", "") or ""
        field  = edu.get("fieldOfStudy", "") or ""

        # Dates are nested under timePeriod → {startDate, endDate}
        tp = edu.get("timePeriod", {})
        start_year = tp.get("startDate", {}).get("year") if tp.get("startDate") else None
        end_year   = tp.get("endDate",   {}).get("year") if tp.get("endDate")   else None

        entries.append({
            "school":     school.strip(),
            "degree":     degree.strip(),
            "field":      field.strip(),
            "start_year": start_year,
            "end_year":   end_year,
        })

    # Sort: entries with no end_year (still enrolled / most recent) first,
    # then by descending end_year
    entries.sort(key=lambda e: e["end_year"] or 9999, reverse=True)
    return entries

def primary_school(education: list[dict]) -> str:
    """Return the name of the most recent school, or empty string."""
    if not education:
        return ""
    return education[0].get("school", "")


# ── Skills parsing ────────────────────────────────────────────────────────────
def parse_skills(raw_profile: dict) -> list[str]:
    """
    Return a flat list of skill name strings.
    linkedin-api exposes skills under the 'skills' key as a list of dicts.
    """
    skills = []
    for s in raw_profile.get("skills", []):
        name = s.get("name", "").strip()
        if name:
            skills.append(name)
    return skills


# ── Excel helpers ────────────────────────────────────────────────────────────
def load_workbook_data() -> tuple[openpyxl.Workbook, dict]:
    """
    Load the master Excel workbook.
    Returns (workbook, {connection_id: row_number_in_snapshot_sheet}).
    """
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"Master workbook not found at {EXCEL_PATH}. "
            "Run pipeline.py at least once first to create it."
        )
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Current Snapshot"]

    # Build a map from connection_id → row number so we can update in place
    id_to_row = {}
    for row in ws.iter_rows(min_row=2):  # row 1 is the header
        cell_id = row[0].value  # column A = connection_id
        if cell_id:
            id_to_row[cell_id] = row[0].row

    return wb, id_to_row


def get_column_index(ws, header_name: str) -> int | None:
    """
    Find the column index (1-based) of a header in row 1.
    Returns None if not found — we add the column in that case.
    """
    for cell in ws[1]:
        if cell.value == header_name:
            return cell.column
    return None


def ensure_columns(ws, new_headers: list[str]) -> dict[str, int]:
    """
    Make sure every header in new_headers exists in row 1.
    Adds missing columns to the right end of the sheet.
    Returns a dict {header_name: column_index (1-based)}.
    """
    header_map = {}
    last_col = ws.max_column

    for h in new_headers:
        idx = get_column_index(ws, h)
        if idx is None:
            last_col += 1
            ws.cell(row=1, column=last_col, value=h)
            idx = last_col
            log.info("Added new column '%s' at column %d", h, idx)
        header_map[h] = idx

    return header_map


def write_enriched_fields(ws, row_num: int, col_map: dict, enriched: dict):
    """Write all enriched fields into the given row of the snapshot sheet."""
    for field, col_idx in col_map.items():
        value = enriched.get(field)
        # Lists are stored as semicolon-separated strings for Excel compatibility
        if isinstance(value, list):
            value = "; ".join(str(v) for v in value) if value else ""
        ws.cell(row=row_num, column=col_idx, value=value)


# ── Connection IDs to enrich ─────────────────────────────────────────────────
def pick_profiles_for_today(id_to_row: dict, wb: openpyxl.Workbook) -> list[str]:
    """
    Select which connection_ids to enrich today.

    Priority order:
      1. Connections that have NEVER been enriched (no school, no skills).
      2. Connections whose enrichment is oldest (based on 'enriched_on' column).

    We pick DAILY_MIN–DAILY_MAX profiles total, shuffled within each tier
    so the ordering feels organic.
    """
    ws = wb["Current Snapshot"]
    enriched_col = get_column_index(ws, "enriched_on")
    school_col   = get_column_index(ws, "school")

    never_enriched = []
    previously_enriched = []  # list of (enriched_on_date, connection_id)

    for conn_id, row_num in id_to_row.items():
        if enriched_col and ws.cell(row=row_num, column=enriched_col).value:
            enriched_date = ws.cell(row=row_num, column=enriched_col).value
            previously_enriched.append((enriched_date, conn_id))
        else:
            never_enriched.append(conn_id)

    # Shuffle never-enriched (avoid always enriching alphabetical order)
    random.shuffle(never_enriched)

    # Sort previously enriched: oldest first so they get refreshed
    previously_enriched.sort(key=lambda x: x[0] or "")
    oldest_ids = [x[1] for x in previously_enriched]

    pool = never_enriched + oldest_ids
    count = random.randint(DAILY_MIN, DAILY_MAX)
    return pool[:count]


# ── LinkedIn URN → profile URL ────────────────────────────────────────────────
def urn_to_public_id(connection_id: str, wb: openpyxl.Workbook, id_to_row: dict) -> str | None:
    """
    Look up the linkedin_public_id (the slug after linkedin.com/in/) stored
    in the snapshot sheet. linkedin-api's get_profile() needs this, not the URN.
    """
    ws = wb["Current Snapshot"]
    col = get_column_index(ws, "linkedin_public_id")
    if col is None:
        return None
    row = id_to_row.get(connection_id)
    if row is None:
        return None
    return ws.cell(row=row, column=col).value


# ── Core enrichment loop ──────────────────────────────────────────────────────
def enrich_profile(api: Linkedin, public_id: str, industry_map: dict) -> dict:
    """
    Fetch a single profile and extract all enriched fields.
    Returns a flat dict ready to be written to Excel.
    """
    log.info("Fetching profile: %s", public_id)
    raw = api.get_profile(public_id)

    # ── Education ────────────────────────────────────────────────────────────
    education = parse_education(raw)
    school = primary_school(education)

    # Flatten education into a readable string for the single 'education' column
    # e.g. "MIT (B.S. Computer Science, 2015–2019); Stanford (M.S. AI, 2019–2021)"
    edu_strings = []
    for e in education:
        parts = []
        if e["degree"]:
            parts.append(e["degree"])
        if e["field"]:
            parts.append(e["field"])
        label = e["school"]
        if parts:
            label += " (" + ", ".join(parts) + ")"
        years = ""
        if e["start_year"] and e["end_year"]:
            years = f", {e['start_year']}–{e['end_year']}"
        elif e["end_year"]:
            years = f", graduated {e['end_year']}"
        label += years
        edu_strings.append(label)
    education_str = "; ".join(edu_strings)

    # ── Current title standardization ────────────────────────────────────────
    # Pull from the first (current) experience entry
    experiences = raw.get("experience", [])
    current_title_raw = ""
    current_company_raw = ""
    if experiences:
        exp = experiences[0]
        current_title_raw   = exp.get("title", "") or ""
        current_company_raw = (exp.get("companyName", "") or
                               exp.get("company", {}).get("companyName", "") or "")

    # Apply full granular taxonomy to the current title
    taxonomy = classify_title(current_title_raw)

    # ── Industry lookup ──────────────────────────────────────────────────────
    industry = lookup_industry(current_company_raw, industry_map)

    # ── Skills ──────────────────────────────────────────────────────────────
    skills = parse_skills(raw)

    # ── Summary ─────────────────────────────────────────────────────────────
    summary = (raw.get("summary", "") or "").strip()

    return {
        "education":            education_str,
        "school":               school,
        # Granular taxonomy (replaces old title_standard)
        "discipline":           taxonomy["discipline"],
        "discipline_family":    taxonomy["discipline_family"],
        "discipline_specialty": taxonomy["discipline_specialty"],
        "seniority":            taxonomy["seniority"],
        # Other enriched fields
        "industry":             industry,
        "skills":               skills,   # stored as semicolon-separated string by writer
        "summary":              summary,
        "enriched_on":          date.today().isoformat(),
    }


def run_enricher():
    log.info("=== Enricher starting ===")
    industry_map = load_industry_map()
    log.info("Industry map loaded: %d entries", len(industry_map))

    wb, id_to_row = load_workbook_data()
    ws_snap = wb["Current Snapshot"]

    # Ensure all new columns exist before we start writing
    NEW_COLUMNS = [
        "education",            # full history string
        "school",               # primary school name only
        "discipline",           # granular discipline e.g. "Electrical — Protection & Relay"
        "discipline_family",    # broad family e.g. "Electrical"
        "discipline_specialty", # specialty e.g. "Protection & Relay"
        "seniority",            # career level e.g. "Senior"
        "industry",             # from industry_map.json
        "skills",               # semicolon-separated skill list
        "summary",              # LinkedIn about/summary section
        "enriched_on",          # ISO date of last enrichment
    ]
    col_map = ensure_columns(ws_snap, NEW_COLUMNS)
    # Save column additions immediately so a crash mid-run doesn't lose them
    wb.save(EXCEL_PATH)

    # Pick today's profiles
    to_enrich = pick_profiles_for_today(id_to_row, wb)
    log.info("Profiles to enrich today: %d", len(to_enrich))

    if not to_enrich:
        log.info("All connections have been enriched recently. Nothing to do.")
        return

    # Authenticate once
    api = Linkedin(config.LINKEDIN_EMAIL, config.LINKEDIN_PASSWORD)
    log.info("LinkedIn auth successful")

    for i, conn_id in enumerate(to_enrich):
        public_id = urn_to_public_id(conn_id, wb, id_to_row)
        if not public_id:
            log.warning("No public_id found for connection %s — skipping", conn_id)
            continue

        try:
            enriched = enrich_profile(api, public_id, industry_map)
            row_num = id_to_row[conn_id]
            write_enriched_fields(ws_snap, row_num, col_map, enriched)
            wb.save(EXCEL_PATH)  # save after each profile so partial progress is kept
            log.info(
                "[%d/%d] Enriched %s — school: %s | industry: %s | title_std: %s | skills: %d",
                i+1, len(to_enrich),
                public_id,
                enriched["school"] or "—",
                enriched["industry"],
                enriched["title_standard"],
                len(enriched["skills"]) if isinstance(enriched["skills"], list) else 0,
            )
        except Exception as e:
            log.error("Failed to enrich %s: %s", public_id, e)
            # Don't crash the whole run — move on to next profile
            continue

        # Randomized delay between fetches
        if i < len(to_enrich) - 1:
            delay = random.randint(DELAY_MIN, DELAY_MAX)
            log.info("Waiting %d seconds before next fetch...", delay)
            time.sleep(delay)

    log.info("=== Enricher complete. %d profiles processed. ===", len(to_enrich))


if __name__ == "__main__":
    run_enricher()
