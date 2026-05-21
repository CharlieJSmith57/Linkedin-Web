"""
export_to_json.py — Static data exporter

Reads network_master.xlsx, applies taxonomy, and writes docs/data/network.json.

Run after pipeline.py or ingest_enriched.py:
    python export_to_json.py

Includes all fields from both the CSV pipeline and the HTML scraper.
"""
from __future__ import annotations

import json
import logging
import re
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

import config
from title_taxonomy import classify

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [EXPORTER] %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

EXCEL_PATH  = Path(config.EXCEL_PATH)
OUTPUT_PATH = Path(config.JSON_OUTPUT_PATH)


# ── Helpers ───────────────────────────────────────────────────────────────────

def sheet_to_dicts(ws) -> list[dict]:
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else f"col_{i}"
               for i, h in enumerate(next(rows))]
    return [
        {headers[i]: cell for i, cell in enumerate(row)}
        for row in rows
        if any(cell is not None for cell in row)
    ]

def safe_str(val) -> str:
    if val is None: return ""
    return str(val).strip()

def safe_date(val) -> str | None:
    if val is None: return None
    if isinstance(val, (datetime, date)):
        return val.date().isoformat() if isinstance(val, datetime) else val.isoformat()
    s = str(val).strip()
    if not s or s.lower() in ("none", "null", ""): return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try: return datetime.strptime(s, fmt).date().isoformat()
        except ValueError: continue
    return s

def split_semicolon(val) -> list[str]:
    if not val: return []
    return [x.strip() for x in str(val).split(";") if x.strip()]

def split_pipe(val: str) -> list[str]:
    if not val: return []
    return [x.strip() for x in str(val).split("|") if x.strip()]

def parse_company_locations(val: str) -> dict:
    """Parse pipe-separated 'Company::Location' into {company: location}."""
    result = {}
    for item in split_pipe(val):
        if "::" in item:
            co, loc = item.split("::", 1)
            result[co.strip()] = loc.strip()
    return result

EDU_RE = re.compile(r"^(?P<school>[^(]+?)(?:\s*\((?P<detail>[^)]+)\))?\s*$")
YEAR_RANGE_RE = re.compile(r"(\d{4})\s*[–\-]\s*(\d{4})")
SINGLE_YEAR_RE = re.compile(r"graduated\s+(\d{4})")

def parse_edu_string(edu_str: str) -> list[dict]:
    entries = []
    for chunk in split_semicolon(edu_str):
        m = EDU_RE.match(chunk)
        if not m:
            entries.append({"school": chunk, "degree": "", "field": "", "start_year": None, "end_year": None})
            continue
        school = m.group("school").strip()
        detail = m.group("detail") or ""
        yr = YEAR_RANGE_RE.search(detail)
        start_year = int(yr.group(1)) if yr else None
        end_year   = int(yr.group(2)) if yr else None
        if not end_year:
            sy = SINGLE_YEAR_RE.search(detail)
            if sy: end_year = int(sy.group(1))
        detail_clean = YEAR_RANGE_RE.sub("", detail)
        detail_clean = SINGLE_YEAR_RE.sub("", detail_clean).strip(" ,")
        parts = [p.strip() for p in detail_clean.split(",", 1) if p.strip()]
        entries.append({
            "school":     school,
            "degree":     parts[0] if parts else "",
            "field":      parts[1] if len(parts) > 1 else "",
            "start_year": start_year,
            "end_year":   end_year,
        })
    return entries

def build_history_map(history_rows: list[dict]) -> dict:
    history_map = {}
    for row in history_rows:
        conn_id = safe_str(row.get("connection_id"))
        if not conn_id: continue
        title   = safe_str(row.get("title") or row.get("job_title") or "")
        company = safe_str(row.get("company") or row.get("company_name") or "")
        start   = safe_date(row.get("start") or row.get("start_date"))
        end     = safe_date(row.get("end")   or row.get("end_date"))
        current = str(row.get("is_current") or "").upper() in ("YES", "TRUE", "1")
        taxonomy = classify(title)
        entry = {
            "company":              company,
            "title":                title,
            "start":                start,
            "end":                  end,
            "current":              current,
            "discipline":           taxonomy["discipline"],
            "discipline_family":    taxonomy["discipline_family"],
            "discipline_specialty": taxonomy["discipline_specialty"],
            "seniority":            taxonomy["seniority"],
            "is_pe":                taxonomy.get("is_pe", False),
        }
        history_map.setdefault(conn_id, []).append(entry)
    for conn_id in history_map:
        history_map[conn_id].sort(
            key=lambda e: (0 if e["current"] else 1, e["start"] or "0000"),
            reverse=False,
        )
        history_map[conn_id] = list(reversed(history_map[conn_id]))
    return history_map


# ── Main export ───────────────────────────────────────────────────────────────

def export():
    log.info("Loading workbook: %s", EXCEL_PATH)
    if not EXCEL_PATH.exists():
        log.error("Workbook not found. Run pipeline.py first.")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    if "Current Snapshot" not in wb.sheetnames:
        log.error("'Current Snapshot' sheet not found.")
        sys.exit(1)

    snapshot_rows = sheet_to_dicts(wb["Current Snapshot"])
    log.info("Snapshot rows: %d", len(snapshot_rows))

    history_rows = []
    if "Employment History" in wb.sheetnames:
        history_rows = sheet_to_dicts(wb["Employment History"])
        log.info("History rows: %d", len(history_rows))

    wb.close()
    history_map = build_history_map(history_rows)

    connections = []
    skipped = 0

    for row in snapshot_rows:
        conn_id = safe_str(row.get("connection_id") or row.get("linkedin_id") or "")
        first   = safe_str(row.get("first_name") or row.get("first") or "")
        last    = safe_str(row.get("last_name")  or row.get("last")  or "")
        if not conn_id or (not first and not last):
            skipped += 1
            continue

        current_title = safe_str(row.get("current_title") or row.get("title") or "")
        taxonomy = classify(current_title)

        # Cert tags — combine taxonomy PE with scraper cert_tags
        scraper_cert_tags = split_pipe(safe_str(row.get("cert_tags") or ""))
        all_cert_tags = list(scraper_cert_tags)
        if taxonomy.get("is_pe") and "PE" not in all_cert_tags:
            all_cert_tags.insert(0, "PE")
        # Custom certifications from taxonomy
        for cert_name, cert_val in taxonomy.get("certifications", {}).items():
            if cert_val and cert_name not in all_cert_tags:
                all_cert_tags.append(cert_name)

        conn = {
            # Identity
            "id":                   conn_id,
            "first":                first,
            "last":                 last,
            "headline":             safe_str(row.get("headline") or ""),
            "location":             safe_str(row.get("location") or ""),
            "photo_url":            safe_str(row.get("profile_photo_url") or row.get("photo_url") or ""),
            "banner_url":           safe_str(row.get("banner_url") or ""),
            "connected_on":         safe_date(row.get("connected_on") or row.get("data_since")),

            # Current role
            "current_company":      safe_str(row.get("current_company") or ""),
            "current_title":        current_title,
            "current_start":        safe_date(row.get("current_start_date") or row.get("current_start")),

            # Taxonomy
            "discipline":           taxonomy["discipline"],
            "discipline_family":    taxonomy["discipline_family"],
            "discipline_specialty": taxonomy["discipline_specialty"],
            "seniority":            taxonomy["seniority"],
            "is_pe":                taxonomy.get("is_pe", False),
            "certifications":       taxonomy.get("certifications", {}),

            # Enriched — API scraper
            "school":               safe_str(row.get("school") or ""),
            "industry":             safe_str(row.get("industry") or "Unknown"),
            "skills":               split_semicolon(safe_str(row.get("skills") or "")),
            "summary":              safe_str(row.get("summary") or ""),
            "enriched_on":          safe_date(row.get("enriched_on")),

            # Enriched — HTML scraper (ingest_enriched.py)
            "cert_tags":            all_cert_tags,
            "greek_orgs":           split_pipe(safe_str(row.get("greek_orgs") or "")),
            "company_locations":    parse_company_locations(safe_str(row.get("company_locations") or "")),
            "honors":               split_pipe(safe_str(row.get("honors") or "")),
            "org_memberships":      split_pipe(safe_str(row.get("org_memberships") or "")),
            "scraper_synced_on":    safe_date(row.get("scraper_synced_on")),

            # Nested
            "education":            parse_edu_string(safe_str(row.get("education") or "")),
            "history":              history_map.get(conn_id, []),
        }
        connections.append(conn)

    connections.sort(key=lambda c: (c["last"].lower(), c["first"].lower()))

    payload = {
        "generated":   datetime.now().isoformat(timespec="seconds"),
        "count":       len(connections),
        "connections": connections,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    log.info("Exported %d connections → %s (skipped %d)", len(connections), OUTPUT_PATH, skipped)

    from collections import Counter
    families = Counter(c["discipline_family"] for c in connections)
    log.info("Discipline breakdown:")
    for family, count in families.most_common():
        log.info("  %-30s %d", family, count)

    # Scraper coverage summary
    scraped = sum(1 for c in connections if c.get("scraper_synced_on"))
    log.info("Scraper coverage: %d / %d profiles enriched (%.0f%%)",
             scraped, len(connections),
             100 * scraped / len(connections) if connections else 0)


if __name__ == "__main__":
    export()
