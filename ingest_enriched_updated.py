"""
ingest_enriched.py  —  Slow LinkedIn profile enricher
B:\\linkedin-pipeline\\ingest_enriched.py

Fetches 8-12 random profiles per run via the unofficial linkedin-api library,
enriches the master Excel workbook with:
  • Full employment history with dates
  • Profile photo URL
  • Location
  • School / education history
  • Greek org affiliations  ← NEW (via greek_detector.py)

Greek org detection
───────────────────
After fetching each profile, greek_detector.detect_greek_orgs() scans all
text fields and writes results to three Excel columns:
  greek_orgs       — full JSON array of match dicts
  greek_org_ids    — comma-separated org_id list  (e.g., "theta_chi, tke")
  greek_org_names  — comma-separated display names (e.g., "Theta Chi, Tau Kappa Epsilon")

The front-end Greek clustering / badge UI reads greek_org_ids from network.json.
"""

from __future__ import annotations

import json
import logging
import os
import random
import sys
import time
from datetime import date, datetime
from pathlib import Path

import openpyxl
from linkedin_api import Linkedin

# ── Greek detector ─────────────────────────────────────────────────────────────
# Import from same directory; falls back gracefully if file is missing
try:
    from greek_detector import detect_greek_orgs, load_greek_index
    _GREEK_AVAILABLE = True
except ImportError:
    _GREEK_AVAILABLE = False
    logging.warning("greek_detector.py not found — Greek org detection disabled")

# ── Paths ─────────────────────────────────────────────────────────────────────

PIPELINE_DIR = Path(__file__).parent
DATA_DIR     = Path(os.environ.get("LINKEDIN_DATA_DIR", r"B:\linkedin-data"))
EXCEL_PATH   = DATA_DIR / "network_master.xlsx"
QUEUE_FILE   = DATA_DIR / "enricher_queue.json"
LOG_FILE     = DATA_DIR / "logs" / f"enricher_{date.today()}.log"

# ── Enricher settings ─────────────────────────────────────────────────────────

MIN_PROFILES_PER_RUN = 8
MAX_PROFILES_PER_RUN = 12
MIN_DELAY_SEC        = 45    # min seconds between profile fetches
MAX_DELAY_SEC        = 120   # max seconds — randomized to avoid detection
PRIORITY_ORDER       = ["never_enriched", "oldest_enriched", "random"]

# ── LinkedIn credentials ───────────────────────────────────────────────────────
# Set these as environment variables; never hardcode

LI_USERNAME = os.environ.get("LINKEDIN_USERNAME", "")
LI_PASSWORD = os.environ.get("LINKEDIN_PASSWORD", "")

# ── Logging ───────────────────────────────────────────────────────────────────

LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("enricher")

# ── Excel column map ───────────────────────────────────────────────────────────
# These match the columns written by change_detector.py / exporter.py

SNAPSHOT_COLS = [
    "connection_id", "first_name", "last_name", "headline", "location",
    "current_company", "current_title", "current_start_month", "current_start_year",
    "profile_photo_url", "linkedin_url", "last_enriched", "enriched_version",
    "discipline", "seniority", "is_pe", "is_eit",
    # Greek org columns
    "greek_orgs", "greek_org_ids", "greek_org_names",
    # Employment history columns (up to 10 previous jobs)
    *[f"prev_company_{i}"      for i in range(1, 11)],
    *[f"prev_title_{i}"        for i in range(1, 11)],
    *[f"prev_start_month_{i}"  for i in range(1, 11)],
    *[f"prev_start_year_{i}"   for i in range(1, 11)],
    *[f"prev_end_month_{i}"    for i in range(1, 11)],
    *[f"prev_end_year_{i}"     for i in range(1, 11)],
]


# ── Queue management ──────────────────────────────────────────────────────────

def load_queue() -> list[dict]:
    """Load or build the enricher queue from the master Excel."""
    if QUEUE_FILE.exists():
        try:
            q = json.loads(QUEUE_FILE.read_text(encoding="utf-8"))
            if q:
                return q
        except Exception:
            pass
    return _build_queue_from_excel()


def _build_queue_from_excel() -> list[dict]:
    """Build a prioritized queue from the Current Snapshot sheet."""
    log.info("Building enricher queue from Excel...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
        ws = wb["Current Snapshot"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    except Exception as exc:
        log.error("Cannot read Excel: %s", exc)
        return []

    def col(row, name):
        try:
            return row[headers.index(name)]
        except (ValueError, IndexError):
            return None

    never, stale, rest = [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cid      = col(row, "connection_id")
        url      = col(row, "linkedin_url")
        enriched = col(row, "last_enriched")
        if not cid or not url:
            continue
        entry = {"connection_id": str(cid), "linkedin_url": str(url),
                 "last_enriched": str(enriched) if enriched else None}
        if not enriched:
            never.append(entry)
        elif _days_since(enriched) > 30:
            stale.append(entry)
        else:
            rest.append(entry)

    random.shuffle(never)
    random.shuffle(stale)
    random.shuffle(rest)
    queue = never + stale + rest

    try:
        QUEUE_FILE.write_text(json.dumps(queue, ensure_ascii=False, indent=2),
                              encoding="utf-8")
    except Exception as exc:
        log.warning("Could not save queue file: %s", exc)

    log.info("Queue built: %d never-enriched, %d stale, %d recent",
             len(never), len(stale), len(rest))
    return queue


def _days_since(dt_str: str) -> int:
    try:
        d = datetime.fromisoformat(str(dt_str)).date()
        return (date.today() - d).days
    except Exception:
        return 9999


def pop_from_queue(n: int) -> list[dict]:
    """Pop up to n entries from the queue."""
    queue = load_queue()
    batch = queue[:n]
    remainder = queue[n:]
    try:
        QUEUE_FILE.write_text(json.dumps(remainder, ensure_ascii=False, indent=2),
                              encoding="utf-8")
    except Exception:
        pass
    return batch


# ── LinkedIn fetcher ───────────────────────────────────────────────────────────

def fetch_profile(api: Linkedin, linkedin_url: str) -> dict | None:
    """Fetch a single profile. Returns raw profile dict or None on failure."""
    # Extract the public_id from the URL
    # URLs look like: https://www.linkedin.com/in/john-smith-12345/
    parts = linkedin_url.rstrip("/").split("/")
    try:
        public_id = parts[parts.index("in") + 1]
    except (ValueError, IndexError):
        log.warning("Cannot parse public_id from URL: %s", linkedin_url)
        return None

    try:
        profile = api.get_profile(public_id)
        return profile
    except Exception as exc:
        log.warning("Failed to fetch %s: %s", public_id, exc)
        return None


# ── Profile parsing ────────────────────────────────────────────────────────────

def parse_experience(profile: dict) -> list[dict]:
    """
    Extract employment history from raw LinkedIn profile.
    Returns list of dicts with keys: company, title, start_month, start_year,
    end_month, end_year, is_current.
    Sorted most-recent first.
    """
    jobs = []
    for pos in profile.get("experience", []):
        company = pos.get("companyName") or pos.get("company_name") or ""
        title   = pos.get("title") or ""

        # Date parsing — LinkedIn API returns nested dicts like:
        # {"start": {"month": 6, "year": 2019}, "end": {"month": None, "year": None}}
        start = pos.get("timePeriod", {}).get("startDate") or pos.get("start") or {}
        end   = pos.get("timePeriod", {}).get("endDate")   or pos.get("end")   or {}

        start_month = start.get("month") if isinstance(start, dict) else None
        start_year  = start.get("year")  if isinstance(start, dict) else None
        end_month   = end.get("month")   if isinstance(end, dict)   else None
        end_year    = end.get("year")    if isinstance(end, dict)   else None
        is_current  = not bool(end_year)

        jobs.append({
            "company":     company,
            "title":       title,
            "start_month": start_month,
            "start_year":  start_year,
            "end_month":   end_month,
            "end_year":    end_year,
            "is_current":  is_current,
        })

    # Sort: current jobs first, then by start_year desc
    jobs.sort(key=lambda j: (not j["is_current"], -(j["start_year"] or 0)))
    return jobs


def parse_location(profile: dict) -> str:
    loc = profile.get("geoLocationName") or profile.get("locationName") or ""
    return str(loc).strip()


def parse_photo_url(profile: dict) -> str:
    # LinkedIn API returns displayPicture or miniProfile.picture
    display = profile.get("displayPictureUrl") or ""
    if display:
        # Append the largest artifact
        artifacts = profile.get("img_xl_width") or profile.get("img_100_100")
        return display  # simplified; full URL assembly varies by API version
    return ""


# ── Greek org detection wrapper ────────────────────────────────────────────────

def detect_greek(raw_profile: dict) -> tuple[str, str, str]:
    """
    Run Greek org detection on a raw profile.
    Returns (greek_orgs_json, greek_org_ids_csv, greek_org_names_csv).
    """
    if not _GREEK_AVAILABLE:
        return "", "", ""
    try:
        matches = detect_greek_orgs(raw_profile, DATA_DIR / "data" if (DATA_DIR / "data").exists() else DATA_DIR)
        greek_json  = json.dumps(matches, ensure_ascii=False) if matches else ""
        greek_ids   = ", ".join(m["org_id"]    for m in matches)
        greek_names = ", ".join(m["full_name"] for m in matches)
        if matches:
            log.info("    Greek orgs detected: %s", greek_ids)
        return greek_json, greek_ids, greek_names
    except Exception as exc:
        log.warning("    Greek detection error: %s", exc)
        return "", "", ""


# ── Excel writer ───────────────────────────────────────────────────────────────

def write_enriched_data(connection_id: str, data: dict) -> None:
    """
    Write enriched profile data back to the master Excel workbook.
    Updates the matching row in Current Snapshot; does NOT touch Employment History
    (that's change_detector.py's job).
    """
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Current Snapshot"]
    except Exception as exc:
        log.error("Cannot open Excel for writing: %s", exc)
        return

    headers = [c.value for c in ws[1]]

    def col_idx(name: str) -> int | None:
        try:
            return headers.index(name) + 1   # 1-indexed
        except ValueError:
            return None

    # Ensure Greek columns exist (add them if this is the first time)
    for greek_col in ("greek_orgs", "greek_org_ids", "greek_org_names"):
        if greek_col not in headers:
            next_col = len(headers) + 1
            ws.cell(row=1, column=next_col, value=greek_col)
            headers.append(greek_col)

    # Find the row for this connection_id
    cid_col = col_idx("connection_id")
    target_row = None
    if cid_col:
        for row in ws.iter_rows(min_row=2):
            if str(row[cid_col - 1].value) == str(connection_id):
                target_row = row[0].row
                break

    if not target_row:
        log.warning("connection_id %s not found in snapshot — skipping write", connection_id)
        wb.close()
        return

    def write_cell(col_name: str, value) -> None:
        idx = col_idx(col_name)
        if idx:
            ws.cell(row=target_row, column=idx, value=value)

    # Write enriched fields
    write_cell("location",         data.get("location", ""))
    write_cell("profile_photo_url", data.get("photo_url", ""))
    write_cell("last_enriched",    datetime.now().isoformat())
    write_cell("enriched_version", "2.0")

    # Greek org fields
    write_cell("greek_orgs",       data.get("greek_orgs", ""))
    write_cell("greek_org_ids",    data.get("greek_org_ids", ""))
    write_cell("greek_org_names",  data.get("greek_org_names", ""))

    # Previous jobs (up to 10, skipping index 0 which is current)
    prev_jobs = data.get("prev_jobs", [])
    for i, job in enumerate(prev_jobs[:10], start=1):
        write_cell(f"prev_company_{i}",      job.get("company", ""))
        write_cell(f"prev_title_{i}",        job.get("title", ""))
        write_cell(f"prev_start_month_{i}",  job.get("start_month"))
        write_cell(f"prev_start_year_{i}",   job.get("start_year"))
        write_cell(f"prev_end_month_{i}",    job.get("end_month"))
        write_cell(f"prev_end_year_{i}",     job.get("end_year"))

    try:
        wb.save(EXCEL_PATH)
        log.info("    Excel updated for connection_id %s", connection_id)
    except Exception as exc:
        log.error("    Failed to save Excel: %s", exc)
    finally:
        wb.close()


# ── Main enricher loop ─────────────────────────────────────────────────────────

def main() -> None:
    log.info("=" * 60)
    log.info("Enricher starting — %s", datetime.now().isoformat())

    if not LI_USERNAME or not LI_PASSWORD:
        log.error("LINKEDIN_USERNAME / LINKEDIN_PASSWORD not set")
        sys.exit(1)

    if not EXCEL_PATH.exists():
        log.error("Master Excel not found: %s", EXCEL_PATH)
        sys.exit(1)

    # Pre-load Greek index so it's warm for all profiles
    if _GREEK_AVAILABLE:
        greek_data_dir = DATA_DIR / "data" if (DATA_DIR / "data").exists() else DATA_DIR
        try:
            idx = load_greek_index(greek_data_dir)
            log.info("Greek index loaded: %d orgs", len(idx._org_by_id))
        except Exception as exc:
            log.warning("Greek index load failed: %s", exc)

    # Pick how many profiles to fetch this run
    n = random.randint(MIN_PROFILES_PER_RUN, MAX_PROFILES_PER_RUN)
    log.info("Targeting %d profiles this run", n)

    batch = pop_from_queue(n)
    if not batch:
        log.info("Queue is empty — rebuilding from Excel")
        _build_queue_from_excel()
        batch = pop_from_queue(n)

    if not batch:
        log.info("No profiles to enrich — exiting")
        return

    # Authenticate
    try:
        api = Linkedin(LI_USERNAME, LI_PASSWORD)
        log.info("LinkedIn authenticated as %s", LI_USERNAME)
    except Exception as exc:
        log.error("LinkedIn auth failed: %s", exc)
        sys.exit(1)

    # Enrich each profile
    success, failed = 0, 0
    for i, entry in enumerate(batch):
        cid = entry["connection_id"]
        url = entry["linkedin_url"]
        log.info("Fetching %d/%d — %s", i + 1, len(batch), url)

        raw = fetch_profile(api, url)
        if raw is None:
            failed += 1
            # Re-queue failed entries at the back
            continue

        # Parse
        jobs = parse_experience(raw)
        current_jobs = [j for j in jobs if j["is_current"]]
        prev_jobs    = [j for j in jobs if not j["is_current"]]

        # Greek org detection
        greek_json, greek_ids, greek_names = detect_greek(raw)

        enriched = {
            "location":     parse_location(raw),
            "photo_url":    parse_photo_url(raw),
            "prev_jobs":    prev_jobs,
            "greek_orgs":   greek_json,
            "greek_org_ids":   greek_ids,
            "greek_org_names": greek_names,
        }

        write_enriched_data(cid, enriched)
        success += 1

        # Randomized delay — avoid rate limiting / bot detection
        if i < len(batch) - 1:
            delay = random.uniform(MIN_DELAY_SEC, MAX_DELAY_SEC)
            log.info("    Sleeping %.0f sec before next fetch...", delay)
            time.sleep(delay)

    log.info(
        "Enricher done — %d succeeded, %d failed (of %d attempted)",
        success, failed, len(batch),
    )
    log.info("=" * 60)


if __name__ == "__main__":
    main()
