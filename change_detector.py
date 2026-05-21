"""
change_detector.py — CSV diff engine.
Reads Connections.csv, diffs against network_master.xlsx, writes changes.
Usage: python change_detector.py path/to/Connections.csv
"""
from __future__ import annotations
import csv, logging, os, sys
from datetime import date, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import config
from title_taxonomy import classify

log = logging.getLogger(__name__)
EXCEL_PATH = Path(config.EXCEL_PATH)
LINKEDIN_BLUE = "0A66C2"
HEADER_FILL = PatternFill("solid", fgColor=LINKEDIN_BLUE)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL = PatternFill("solid", fgColor="F3F6FA")

SNAP_COLS = ["connection_id","linkedin_public_id","first_name","last_name","headline",
    "location","current_company","current_title","current_start_date","profile_photo_url",
    "school","discipline","discipline_family","discipline_specialty","seniority","is_pe",
    "industry","skills","education","summary","enriched_on","connected_on","data_since",
    "last_updated","change_log"]
HIST_COLS = ["connection_id","first_name","last_name","company","title","discipline",
    "discipline_family","discipline_specialty","seniority","start","end","is_current",
    "change_type","recorded_on"]

def parse_csv(csv_path):
    records = []
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        lines = f.readlines()
    header_idx = next((i for i,l in enumerate(lines) if "first" in l.lower() and "name" in l.lower()), 0)
    reader = csv.DictReader("".join(lines[header_idx:]).splitlines())
    for row in reader:
        norm = {k.strip().lower().replace(" ","_"):(v or "").strip() for k,v in row.items()}
        first = norm.get("first_name") or norm.get("firstname") or ""
        last  = norm.get("last_name")  or norm.get("lastname")  or ""
        url   = norm.get("url") or norm.get("profile_url") or ""
        company  = norm.get("company") or norm.get("current_company") or ""
        position = norm.get("position") or norm.get("title") or ""
        connected = norm.get("connected_on") or ""
        if not first and not last: continue
        slug = url.split("/in/")[-1].split("?")[0].strip("/") if "/in/" in url else url.replace("https://","").replace("www.linkedin.com/in/","").strip("/")
        conn_id = "csv_"+slug if slug else f"csv_{first}_{last}".lower().replace(" ","_")
        records.append({"connection_id":conn_id,"linkedin_public_id":slug,"first_name":first,
            "last_name":last,"current_company":company,"current_title":position,"connected_on":connected,"url":url})
    log.info("CSV: %d connections", len(records))
    return records

def load_or_create_workbook():
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames: del wb["Sheet"]
    for name, cols in [("Current Snapshot",SNAP_COLS),("Employment History",HIST_COLS)]:
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            for i,c in enumerate(cols,1):
                cell = ws.cell(row=1,column=i,value=c)
                cell.font=HEADER_FONT; cell.fill=HEADER_FILL
                cell.alignment=Alignment(horizontal="center")
            ws.freeze_panes="A2"
    snap = {}
    ws = wb["Current Snapshot"]
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2,values_only=True):
        if not any(row): continue
        d = {headers[i]:row[i] for i in range(min(len(headers),len(row)))}
        cid = d.get("connection_id")
        if cid: snap[cid] = d
    return wb, snap

def append_row(ws, headers, data):
    rn = ws.max_row+1
    fill = ALT_FILL if rn%2==0 else None
    for i,col in enumerate(headers,1):
        cell = ws.cell(row=rn,column=i,value=data.get(col))
        if fill: cell.fill=fill

def close_history(ws, headers, conn_id, end_date):
    idx = {h:i+1 for i,h in enumerate(headers)}
    cid_col,cur_col,end_col = idx.get("connection_id",1),idx.get("is_current"),idx.get("end")
    for row in ws.iter_rows(min_row=2):
        if row[cid_col-1].value==conn_id and cur_col and row[cur_col-1].value=="YES":
            row[cur_col-1].value="NO"
            if end_col: row[end_col-1].value=end_date

def process_changes(csv_records, snap, wb):
    ws_snap=wb["Current Snapshot"]; ws_hist=wb["Employment History"]
    snap_headers=[c.value for c in ws_snap[1]]; hist_headers=[c.value for c in ws_hist[1]]
    today=date.today().isoformat(); yesterday=(date.today()-timedelta(days=1)).isoformat()
    snap_rows={}
    for row in ws_snap.iter_rows(min_row=2):
        if row[0].value: snap_rows[row[0].value]=row[0].row
    changed=0
    for rec in csv_records:
        cid=rec["connection_id"]; first=rec["first_name"]; last=rec["last_name"]
        new_co=rec["current_company"]; new_title=rec["current_title"]
        tax=classify(new_title)
        if cid not in snap:
            log.info("INITIAL: %s %s @ %s",first,last,new_co)
            row_data={col:None for col in SNAP_COLS}
            row_data.update({"connection_id":cid,"linkedin_public_id":rec["linkedin_public_id"],
                "first_name":first,"last_name":last,"current_company":new_co,"current_title":new_title,
                "current_start_date":today,"discipline":tax["discipline"],"discipline_family":tax["discipline_family"],
                "discipline_specialty":tax["discipline_specialty"],"seniority":tax["seniority"],
                "is_pe":tax.get("is_pe",False),"connected_on":rec.get("connected_on",""),
                "data_since":today,"last_updated":today,"change_log":f"{today}: INITIAL"})
            append_row(ws_snap,SNAP_COLS,row_data)
            if new_co:
                append_row(ws_hist,HIST_COLS,{"connection_id":cid,"first_name":first,"last_name":last,
                    "company":new_co,"title":new_title,"discipline":tax["discipline"],
                    "discipline_family":tax["discipline_family"],"discipline_specialty":tax["discipline_specialty"],
                    "seniority":tax["seniority"],"start":today,"end":None,"is_current":"YES",
                    "change_type":"INITIAL","recorded_on":today})
            changed+=1
        else:
            stored=snap[cid]; old_co=stored.get("current_company") or ""; old_title=stored.get("current_title") or ""
            row_num=snap_rows.get(cid); col_map={h:i+1 for i,h in enumerate(snap_headers)}
            def update(field,val):
                col=col_map.get(field)
                if col and val: ws_snap.cell(row=row_num,column=col,value=val)
            if new_co and new_co!=old_co:
                log.info("COMPANY: %s %s  %s → %s",first,last,old_co,new_co)
                close_history(ws_hist,hist_headers,cid,yesterday)
                append_row(ws_hist,HIST_COLS,{"connection_id":cid,"first_name":first,"last_name":last,
                    "company":new_co,"title":new_title,"discipline":tax["discipline"],
                    "discipline_family":tax["discipline_family"],"discipline_specialty":tax["discipline_specialty"],
                    "seniority":tax["seniority"],"start":today,"end":None,"is_current":"YES",
                    "change_type":"ARRIVAL","recorded_on":today})
                update("current_company",new_co); update("current_title",new_title)
                update("current_start_date",today); update("last_updated",today)
                update("discipline",tax["discipline"]); update("seniority",tax["seniority"])
                changed+=1
            elif new_title and new_title!=old_title:
                log.info("TITLE: %s %s  '%s' → '%s'",first,last,old_title,new_title)
                update("current_title",new_title); update("last_updated",today)
                update("discipline",tax["discipline"]); update("seniority",tax["seniority"])
                changed+=1
    log.info("Changes: %d",changed)
    return changed

def run(csv_path=None):
    if csv_path is None:
        env=os.environ.get("PIPELINE_CSV_PATH")
        if env: csv_path=Path(env)
        elif len(sys.argv)>1: csv_path=Path(sys.argv[1])
        else: log.error("No CSV path. Usage: python change_detector.py Connections.csv"); sys.exit(1)
    if not Path(csv_path).exists(): log.error("CSV not found: %s",csv_path); sys.exit(1)
    records=parse_csv(csv_path)
    if not records: log.warning("No records in CSV"); return
    wb,snap=load_or_create_workbook()
    process_changes(records,snap,wb)
    EXCEL_PATH.parent.mkdir(parents=True,exist_ok=True)
    wb.save(EXCEL_PATH); log.info("Saved: %s",EXCEL_PATH)

if __name__=="__main__":
    logging.basicConfig(level=logging.INFO,format="%(asctime)s [CHANGE] %(levelname)s %(message)s")
    run()
